import pdfplumber
from openpyxl import Workbook
from core.utils.timing import time_it

def merge_side_by_side_tables(tables_with_bbox):
    """
    Groups and merges tables that are side-by-side (overlapping y-coordinates).
    Also orders them from left to right (x-coordinates) and merges row contents.
    """
    if not tables_with_bbox:
        return []
        
    # Sort by y0 first, then x0
    tables_with_bbox.sort(key=lambda x: (x['bbox'][1], x['bbox'][0]))
    
    merged_tables = []
    current_group = []
    
    for t in tables_with_bbox:
        if not current_group:
            current_group.append(t)
        else:
            prev = current_group[-1]
            # If y0 is very close (within 5 points), they are side-by-side
            if abs(t['bbox'][1] - prev['bbox'][1]) < 5:
                current_group.append(t)
            else:
                merged_tables.append(current_group)
                current_group = [t]
    if current_group:
        merged_tables.append(current_group)
        
    final_tables = []
    for group in merged_tables:
        # Sort by x0 (left-to-right)
        group.sort(key=lambda x: x['bbox'][0])
        
        max_rows = max(len(t['data']) for t in group)
        merged_rows = []
        for r_idx in range(max_rows):
            new_row = []
            for t in group:
                if r_idx < len(t['data']):
                    new_row.extend(t['data'][r_idx])
                else:
                    # Pad with empty cells if one table has fewer rows
                    new_row.extend([''] * len(t['data'][0]))
            merged_rows.append(new_row)
        final_tables.append(merged_rows)
        
    return final_tables

@time_it("PDF to Excel (Digital)")
def extract_borderless_table(page):
    """
    Extracts a borderless table from a pdfplumber page using word extraction and gap clustering.
    Returns a list of lists representing the table rows.
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    if not words:
        return []
        
    # 1. Group words into lines based on vertical overlap
    words.sort(key=lambda w: (w['top'], w['x0']))
    lines = []
    current_line = []
    
    for w in words:
        if not current_line:
            current_line.append(w)
        else:
            prev_w = current_line[-1]
            wc_y = (w['top'] + w['bottom']) / 2
            prev_wc_y = (prev_w['top'] + prev_w['bottom']) / 2
            if abs(wc_y - prev_wc_y) < 5:
                current_line.append(w)
            else:
                lines.append(current_line)
                current_line = [w]
    if current_line:
        lines.append(current_line)
        
    # 2. Cluster words into blocks based on horizontal gaps > 15 pixels
    gap_threshold = 15.0
    blocks = []
    for line in lines:
        row_blocks = []
        current_block = [line[0]]
        for i in range(1, len(line)):
            w = line[i]
            prev_w = line[i-1]
            gap = w['x0'] - prev_w['x1']
            if gap > gap_threshold:
                row_blocks.append(current_block)
                current_block = [w]
            else:
                current_block.append(w)
        if current_block:
            row_blocks.append(current_block)
        blocks.append(row_blocks)
        
    # 3. Detect global columns by analyzing block start coordinates (x0)
    block_x0s = []
    for row in blocks:
        for b in row:
            block_x0s.append(b[0]['x0'])
            
    block_x0s.sort()
    global_cols = []
    if block_x0s:
        current_col_x0 = [block_x0s[0]]
        for x in block_x0s[1:]:
            if x - current_col_x0[-1] < 20: # 20 pixels tolerance for column start alignment
                current_col_x0.append(x)
            else:
                global_cols.append(sum(current_col_x0) / len(current_col_x0))
                current_col_x0 = [x]
        if current_col_x0:
            global_cols.append(sum(current_col_x0) / len(current_col_x0))
            
    if not global_cols:
        return []
        
    # 4. Map blocks to columns
    final_table = []
    for row_blocks in blocks:
        row_data = [""] * len(global_cols)
        for b in row_blocks:
            x0 = b[0]['x0']
            closest_col_idx = min(range(len(global_cols)), key=lambda i: abs(global_cols[i] - x0))
            text = " ".join([w['text'] for w in b])
            
            if row_data[closest_col_idx]:
                row_data[closest_col_idx] += " " + text
            else:
                row_data[closest_col_idx] = text
        final_table.append(row_data)
        
    return final_table

def convert_template_to_excel(pdf_path: str, excel_path: str, template: dict, pages: list[int] = None):
    import pdfplumber
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablo_Birlesik"
    has_written = False

    table_settings = {
        "vertical_strategy": "explicit",
        "horizontal_strategy": "explicit",
        "explicit_vertical_lines": template.get("explicit_vertical_lines", []),
        "explicit_horizontal_lines": template.get("explicit_horizontal_lines", []),
    }

    with pdfplumber.open(pdf_path) as pdf:
        target_pages = pages if pages is not None else range(len(pdf.pages))
        for p_idx in target_pages:
            page = pdf.pages[p_idx]
            # Use explicit settings for template
            tables = page.extract_tables(table_settings)
            
            for table in tables:
                if has_written:
                    ws.append([])
                    ws.append([])
                    
                for row in table:
                    cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                    ws.append(cleaned_row)
                has_written = True

    if not has_written:
        ws.append(["Tablo bulunamadı (Şablon eşleşmedi)."])
        
    wb.save(excel_path)

def convert_digital_pdf_to_excel(pdf_path: str, excel_path: str, pages: list[int] = None):
    """
    Converts a digital PDF (or specific pages) to an Excel document.
    Groups all tables on all pages into a single worksheet, merging side-by-side splits.
    """
    import camelot
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablo_Birlesik"
    has_written = False

    pages_str = 'all' if pages is None else ','.join(str(p + 1) for p in pages)
    
    # 1. Try lattice
    tables = None
    try:
        tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice', line_scale=40)
    except Exception as e:
        print(f"Camelot lattice error: {e}")
        
    # If lattice returns nothing, try stream
    if not tables or len(tables) == 0:
        try:
            tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='stream', row_tol=10)
        except Exception as e:
            print(f"Camelot stream error: {e}")

    if tables and len(tables) > 0:
        wb.remove(ws) # Varsayılan boş sekmeyi sil
        for i, table in enumerate(tables):
            sheet_name = f"Tablo_{i+1}"
            new_ws = wb.create_sheet(title=sheet_name)
            for row in table.df.itertuples(index=False, name=None):
                cleaned_row = [str(cell).strip() if cell is not None and str(cell) != "NaN" else "" for cell in row]
                new_ws.append(cleaned_row)
            has_written = True
                    
    # If no tables found at all, write a placeholder message
    if not has_written:
        ws.append(["Tablo bulunamadı."])
        
    wb.save(excel_path)

@time_it("PDF to Excel (OCR)")
def convert_scanned_pdf_to_excel(pdf_path: str, excel_path: str, pages: list[int] = None):
    """
    Converts scanned PDF pages to an Excel document using img2table and RapidOCR.
    Appends all tables from all pages to a single worksheet, preserving merged cells.
    """
    from core.pdf_backend import open_document, render_page, page_count
    from core.ocr.table import extract_tables_to_excel
    import tempfile
    import os
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tablo_Birlesik"
    
    has_written = False
    pdf_doc = open_document(pdf_path)
    
    try:
        total_pages = page_count(pdf_doc)
        if pages is None:
            pages = list(range(total_pages))
            
        with tempfile.TemporaryDirectory() as tmpdir:
            for p_idx in pages:
                img = render_page(pdf_doc, p_idx, dpi=300)
                tmp_xlsx = os.path.join(tmpdir, f"page_{p_idx}.xlsx")
                
                # Use img2table to export directly to a temporary xlsx
                extract_tables_to_excel(img, tmp_xlsx)
                
                if os.path.exists(tmp_xlsx):
                    tmp_wb = openpyxl.load_workbook(tmp_xlsx)
                    for sheet_name in tmp_wb.sheetnames:
                        tmp_ws = tmp_wb[sheet_name]
                        if tmp_ws.max_row > 0:
                            if not has_written:
                                wb.remove(ws) # İlk geçerli tabloda varsayılan sekmeyi sil
                            
                            new_ws_name = f"Sayfa{p_idx+1}_{sheet_name}"
                            new_ws_name = new_ws_name[:31] # Excel ad limiti
                            new_ws = wb.create_sheet(title=new_ws_name)
                            
                            # Add values
                            for row in tmp_ws.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                                        
                            # Copy merged cells
                            for merge in tmp_ws.merged_cells.ranges:
                                new_ws.merge_cells(
                                    start_row=merge.min_row,
                                    start_column=merge.min_col,
                                    end_row=merge.max_row,
                                    end_column=merge.max_col
                                )
                            has_written = True
                    tmp_wb.close()
                    
        if not has_written:
            ws.append(["Tablo bulunamadı."])
            
        wb.save(excel_path)
    finally:
        pdf_doc.close()
